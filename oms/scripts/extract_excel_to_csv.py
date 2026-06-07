#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


DOMAIN_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = DOMAIN_DIR / "docs" / "202604工资明细.xlsx"
ONTOLOGY_PATH = DOMAIN_DIR / "ontology.yaml"
DATA_DIR = DOMAIN_DIR / "data"

PAYROLL_RUN_ID = "PAYROLL_202604"
SNAPSHOT_ID = "SNAPSHOT_202604"
PAYROLL_PERIOD = "2026-04"
ATTENDANCE_PERIOD = "2026-04"
PERFORMANCE_PERIOD = "2026-04"
PAY_DATE = "2026-05-26"
PAYMENT_PERIOD = "2026-05"
CONTRIBUTION_PERIOD = "2026-05"
TAX_PERIOD = "2026-05"


GENERATED_OUTPUT_OBJECTS = {
    "PayrollLine",
    "PayrollItem",
    "SocialInsuranceContribution",
    "HousingFundContribution",
    "ContributionDeduction",
    "Payslip",
    "CostEntry",
    "ConsultantFeeSettlement",
    "PayrollValidationResult",
    "PayrollExport",
    "ApprovalRecord",
}

BENCHMARK_OBJECTS = GENERATED_OUTPUT_OBJECTS | {"TaxLedger"}


COMPANY_ALIASES = {
    "尚诚能源": ("COMP_SCNY", "尚诚能源"),
    "江苏尚诚能源科技有限公司": ("COMP_SCNY", "尚诚能源"),
    "南大尚诚": ("COMP_NDSC", "南大尚诚"),
    "江苏南大尚诚高科技有限公司": ("COMP_NDSC", "南大尚诚"),
    "江苏南大尚诚高科技有限公司（尚诚能源部分）": ("COMP_NDSC", "南大尚诚"),
    "江苏南大尚诚高科技实业有限公司": ("COMP_NDSC", "南大尚诚"),
    "基石": ("COMP_JS", "基石"),
    "南京基石数据技术有限责任公司": ("COMP_JS", "基石"),
}


DETAIL_SHEETS = {
    "COMP_SCNY": "202604工资明细表（尚诚能源）",
    "COMP_NDSC": "202604工资明细表（南大尚诚）",
    "COMP_JS": "202604工资明细表（基石）",
}

SOCIAL_SHEETS = {
    "COMP_SCNY": "202604社保 (尚诚能源)",
    "COMP_NDSC": "202604社保(南大尚诚)",
    "COMP_JS": "202604社保(基石)",
}

HOUSING_SHEETS = {
    "COMP_SCNY": "202604公积金(尚诚能源)",
    "COMP_NDSC": "202604公积金(南大尚诚)",
    "COMP_JS": "202604公积金(基石)",
}

TAX_SHEETS = {
    "COMP_SCNY": "202604综合所得申报税款计算(尚诚能源)",
    "COMP_NDSC": "202604综合所得申报税款计算(南大尚诚)",
    "COMP_JS": "202604综合所得申报税款计算(基石)",
}


def main() -> int:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else WORKBOOK_PATH
    data_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else DATA_DIR
    objects = _load_object_properties()
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    records = {object_type: [] for object_type in objects}
    employees_by_name = _extract_master_payroll(wb, records)
    _extract_company_detail_sheets(wb, records, employees_by_name)
    _extract_social_security(wb, records, employees_by_name)
    _extract_housing_fund(wb, records, employees_by_name)
    _extract_tax_ledgers(wb, records, employees_by_name)
    _extract_payslips(wb, records, employees_by_name)
    _extract_cost_entries(wb, records, employees_by_name)
    _extract_consultants(wb, records)
    _seed_rules(records)
    _write_benchmark_csvs(data_dir / "benchmarks", objects, records)
    _convert_tax_ledgers_to_historical_inputs(records)
    _clear_generated_outputs(records)
    _write_csvs(data_dir, objects, records)

    print(f"wrote OMS CSV data to {data_dir}")
    for object_type in sorted(records):
        print(f"{object_type}: {len(records[object_type])}")
    print(f"wrote Excel benchmark CSV data to {data_dir / 'benchmarks'}")
    return 0


def _load_object_properties() -> dict[str, list[str]]:
    raw = yaml.safe_load(ONTOLOGY_PATH.read_text(encoding="utf-8"))
    return {
        object_type: list(defn.get("properties", {}).keys())
        for object_type, defn in raw.get("objects", {}).items()
    }


def _extract_master_payroll(wb, records: dict[str, list[dict]]) -> dict[str, dict]:
    ws = wb["202604工资表"]
    rows = list(_rows(ws, 2))
    company_ids = {}
    affiliation_ids = {}
    employees_by_name = {}

    records["PayrollRun"].append({
        "payroll_run_id": PAYROLL_RUN_ID,
        "payroll_period": PAYROLL_PERIOD,
        "attendance_period": ATTENDANCE_PERIOD,
        "performance_period": PERFORMANCE_PERIOD,
        "pay_date": PAY_DATE,
        "payment_period": PAYMENT_PERIOD,
        "contribution_period": CONTRIBUTION_PERIOD,
        "tax_period": TAX_PERIOD,
        "status": "draft",
        "source_workbook": "docs/202604工资明细.xlsx",
        "created_by": "excel_extract",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    })
    records["PayrollInputSnapshot"].append({
        "snapshot_id": SNAPSHOT_ID,
        "payroll_run_id": PAYROLL_RUN_ID,
        "payroll_period": PAYROLL_PERIOD,
        "snapshot_time": datetime.now().isoformat(timespec="seconds"),
        "source_note": "由 mock workbook 抽取生成的薪资输入快照",
        "status": "built",
    })

    for index, row in enumerate(rows, start=1):
        name = _str(row.get("姓名"))
        if not name:
            continue
        employee_id = f"EMP{index:03d}"
        raw_company = _str(row.get("公司"))
        company_id, company_name = _company(raw_company)
        affiliation_id = _affiliation_id(raw_company, company_id)
        company_ids[company_id] = company_name
        if affiliation_id:
            affiliation_ids[affiliation_id] = (company_id, raw_company)

        person_id = f"PER{index:03d}"
        employee = {
            "employee_id": employee_id,
            "person_id": person_id,
            "employee_number": "",
            "department": "",
            "position": _str(row.get("岗位")),
            "employment_type": _str(row.get("岗位")),
            "hire_date": "",
            "salary_start_date": _date(row.get("起薪时间")),
            "termination_date": "",
            "status": "在职",
        }
        records["Person"].append({
            "person_id": person_id,
            "name": name,
            "id_type": "",
            "id_number": "",
            "phone": "",
            "bank_account": "",
            "status": "有效",
        })
        records["Employee"].append(employee)
        employees_by_name.setdefault(name, employee)

        employment_relationship_id = _employment_relationship_id(employee)
        records["EmploymentRelationship"].append({
            "employment_relationship_id": employment_relationship_id,
            "employee_id": employee_id,
            "person_id": person_id,
            "effective_date": _date(row.get("起薪时间")) or f"{PAYROLL_PERIOD}-01",
            "expiry_date": "",
            "company_id": company_id,
            "internal_affiliation_id": affiliation_id,
            "relationship_type": _str(row.get("岗位")) or "员工",
            "relationship_reason": "导入初始化",
            "source": "202604工资表",
            "notes": "由 mock workbook 总工资表生成的初始任职关系记录",
        })

        salary_profile_id = f"SAL_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}"
        records["SalaryProfile"].append({
            "salary_profile_id": salary_profile_id,
            "employee_id": employee_id,
            "effective_date": _date(row.get("起薪时间")),
            "expiry_date": "",
            "roster_salary": _num(row.get("花名册薪资")),
            "monthly_salary_base": _num(row.get("月薪基数")),
            "social_security_base": "",
            "housing_fund_base": "",
            "position_or_type": _str(row.get("岗位")),
        })

        employee_snapshot_id = _employee_snapshot_id(employee_id)
        records["PayrollEmployeeSnapshot"].append({
            "employee_snapshot_id": employee_snapshot_id,
            "snapshot_id": SNAPSHOT_ID,
            "payroll_run_id": PAYROLL_RUN_ID,
            "employee_id": employee_id,
            "person_id": person_id,
            "employee_name_snapshot": name,
            "employment_relationship_id": employment_relationship_id,
            "payroll_company_id": company_id,
            "payroll_company_name_snapshot": company_name,
            "internal_affiliation_id": affiliation_id,
            "internal_affiliation_name_snapshot": raw_company if affiliation_id else "",
            "salary_profile_id": salary_profile_id,
            "position_snapshot": _str(row.get("岗位")),
            "employment_status_snapshot": "在职",
            "monthly_salary_base": _num(row.get("月薪基数")),
            "monthly_salary_total": _num(row.get("月薪资总额")),
            "social_security_base": "",
            "housing_fund_base": "",
        })

        payroll_line_id = f"PL_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}"
        attendance_adjustment = ""
        other_adjustment = _num(row.get("补发扣"))
        records["PayrollLine"].append({
            "payroll_line_id": payroll_line_id,
            "payroll_run_id": PAYROLL_RUN_ID,
            "employee_snapshot_id": employee_snapshot_id,
            "employee_id": employee_id,
            "person_id": person_id,
            "employee_name_snapshot": name,
            "employment_relationship_id": employment_relationship_id,
            "payroll_company_id": company_id,
            "payroll_company_name_snapshot": company_name,
            "internal_affiliation_id": affiliation_id,
            "internal_affiliation_name_snapshot": raw_company if affiliation_id else "",
            "salary_profile_id": salary_profile_id,
            "monthly_salary_base": _num(row.get("月薪基数")),
            "monthly_salary_total": _num(row.get("月薪资总额")),
            "position": _str(row.get("岗位")),
            "basic_salary": _num(row.get("基本薪资")),
            "performance_salary_base": _num(row.get("绩效薪资基数")),
            "performance_grade": _str(row.get("绩效等级")),
            "performance_salary": _num(row.get("绩效薪资")),
            "overtime_salary": "",
            "attendance_adjustment": attendance_adjustment,
            "other_adjustment": other_adjustment,
            "gross_pay_before_deduction": _num(row.get("实际应发薪资")),
            "personal_social_security": "",
            "personal_housing_fund": "",
            "personal_income_tax": "",
            "net_pay": "",
            "employer_social_security": "",
            "employer_housing_fund": "",
            "company_total_cost": "",
            "exception_notes": "",
        })

        records["PerformanceRecord"].append({
            "performance_record_id": f"PERF_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}",
            "employee_id": employee_id,
            "performance_period": PERFORMANCE_PERIOD,
            "performance_grade": _str(row.get("绩效等级")),
            "source": "202604工资表",
            "notes": "",
        })

        if other_adjustment not in ("", 0, 0.0):
            records["PayrollAdjustment"].append({
                "adjustment_id": f"ADJ_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}_OTHER",
                "payroll_run_id": PAYROLL_RUN_ID,
                "employee_id": employee_id,
                "adjustment_type": "其他",
                "amount": other_adjustment,
                "reason": "Excel总工资表补发扣",
                "source": "202604工资表",
            })

    for company_id, company_name in company_ids.items():
        records["Company"].append({
            "company_id": company_id,
            "company_name": company_name,
            "payroll_enabled": 1,
            "tax_entity_name": company_name,
            "social_security_account": "",
            "housing_fund_account": "",
        })

    for affiliation_id, (company_id, name) in affiliation_ids.items():
        records["InternalAffiliation"].append({
            "affiliation_id": affiliation_id,
            "company_id": company_id,
            "name": name,
            "description": "由 Excel 公司字段拆分生成",
        })

    return employees_by_name


def _extract_company_detail_sheets(wb, records, employees_by_name):
    payroll_lines = {row["employee_id"]: row for row in records["PayrollLine"]}
    for company_id, sheet_name in DETAIL_SHEETS.items():
        ws = wb[sheet_name]
        company_employees = _employees_for_company(records, company_id)
        for row_offset, row in enumerate(_rows(ws, 2)):
            employee = _mock_aligned_employee(
                records,
                employees_by_name,
                _str(row.get("姓名")),
                company_id=company_id,
                source=sheet_name,
                row_offset=row_offset,
                company_employees=company_employees,
            )
            if not employee:
                continue
            line = payroll_lines.get(employee["employee_id"])
            if not line:
                _append_mock_warning(
                    records,
                    target_type="PayrollLine",
                    target_id=employee["employee_id"],
                    rule_code="mock_payroll_line_missing",
                    message=f"{sheet_name} 第 {row.get('_row_index')} 行已对齐到 {employee['employee_id']}，但总工资表没有对应工资明细",
                )
                continue
            line["personal_social_security"] = _num(row.get("社保合计"))
            line["personal_housing_fund"] = _num(row.get("住房公积金"))
            line["personal_income_tax"] = _num(row.get("本月个税"))
            line["net_pay"] = _num(row.get("实发工资"))
            if _as_float(line["net_pay"]) < 0:
                line["exception_notes"] = "实发工资为负数，请确认扣款是否正确"


def _extract_social_security(wb, records, employees_by_name):
    for company_id, sheet_name in SOCIAL_SHEETS.items():
        ws = wb[sheet_name]
        company_name = _str(ws.cell(2, 3).value) or _company_name(company_id)
        _ensure_company(records, company_id, company_name)
        company_employees = _employees_for_company(records, company_id)
        row_offset = 0
        for row_index in range(5, ws.max_row + 1):
            name = _str(ws.cell(row_index, 3).value)
            if not name:
                continue
            employee = _mock_aligned_employee(
                records,
                employees_by_name,
                name,
                company_id=company_id,
                source=sheet_name,
                row_offset=row_offset,
                company_employees=company_employees,
            )
            row_offset += 1
            if not employee:
                continue
            contribution_id = f"SI_{CONTRIBUTION_PERIOD.replace('-', '')}_{company_id}_{employee['employee_id']}"
            employer_total = _num(ws.cell(row_index, 10).value)
            personal_total = _num(ws.cell(row_index, 15).value)
            records["SocialInsuranceContribution"].append({
                "contribution_id": contribution_id,
                "employee_id": employee["employee_id"],
                "contribution_company_id": company_id,
                "contribution_period": CONTRIBUTION_PERIOD,
                "contribution_base": _num(ws.cell(row_index, 4).value),
                "employer_pension": _num(ws.cell(row_index, 5).value),
                "employer_injury": _num(ws.cell(row_index, 6).value),
                "employer_unemployment": _num(ws.cell(row_index, 7).value),
                "employer_medical": _num(ws.cell(row_index, 8).value),
                "employer_maternity": _num(ws.cell(row_index, 9).value),
                "employer_total": employer_total,
                "personal_pension": _num(ws.cell(row_index, 11).value),
                "personal_unemployment": _num(ws.cell(row_index, 12).value),
                "personal_medical": _num(ws.cell(row_index, 13).value),
                "personal_serious_illness": _num(ws.cell(row_index, 14).value),
                "personal_total": personal_total,
                "status": "deducted",
            })
            records["ContributionDeduction"].append({
                "deduction_id": f"DED_SI_{CONTRIBUTION_PERIOD.replace('-', '')}_{company_id}_{employee['employee_id']}",
                "employee_id": employee["employee_id"],
                "person_id": employee.get("person_id", ""),
                "employee_snapshot_id": _employee_snapshot_id(employee["employee_id"]),
                "employment_relationship_id": _employment_relationship_id(employee),
                "deduction_type": "社保",
                "contribution_record_id": contribution_id,
                "contribution_company_id": company_id,
                "contribution_period": CONTRIBUTION_PERIOD,
                "personal_amount": personal_total,
                "employer_amount": employer_total,
                "deducted_payroll_run_id": PAYROLL_RUN_ID,
                "deducted_payroll_period": PAYROLL_PERIOD,
                "deduction_reason": "正常扣款",
                "status": "deducted",
            })


def _extract_housing_fund(wb, records, employees_by_name):
    for company_id, sheet_name in HOUSING_SHEETS.items():
        ws = wb[sheet_name]
        company_employees = _employees_for_company(records, company_id)
        for row_offset, row in enumerate(_rows(ws, 4)):
            employee = _mock_aligned_employee(
                records,
                employees_by_name,
                _str(row.get("姓名")),
                company_id=company_id,
                source=sheet_name,
                row_offset=row_offset,
                company_employees=company_employees,
            )
            if not employee:
                continue
            housing_fund_id = f"HF_{CONTRIBUTION_PERIOD.replace('-', '')}_{company_id}_{employee['employee_id']}"
            employer_amount = _num_by_prefix(row, "单位")
            personal_amount = _num_by_prefix(row, "个人")
            records["HousingFundContribution"].append({
                "housing_fund_id": housing_fund_id,
                "employee_id": employee["employee_id"],
                "contribution_company_id": company_id,
                "contribution_period": CONTRIBUTION_PERIOD,
                "contribution_base": _num(row.get("基数")),
                "employer_rate": 0.12 if company_id == "COMP_JS" else 0.10,
                "personal_rate": 0.12 if company_id == "COMP_JS" else 0.10,
                "employer_amount": employer_amount,
                "personal_amount": personal_amount,
                "total_amount": _num(row.get("合计")),
                "rounding_rule": "ROUND0",
                "status": "deducted",
            })
            records["ContributionDeduction"].append({
                "deduction_id": f"DED_HF_{CONTRIBUTION_PERIOD.replace('-', '')}_{company_id}_{employee['employee_id']}",
                "employee_id": employee["employee_id"],
                "person_id": employee.get("person_id", ""),
                "employee_snapshot_id": _employee_snapshot_id(employee["employee_id"]),
                "employment_relationship_id": _employment_relationship_id(employee),
                "deduction_type": "公积金",
                "contribution_record_id": housing_fund_id,
                "contribution_company_id": company_id,
                "contribution_period": CONTRIBUTION_PERIOD,
                "personal_amount": personal_amount,
                "employer_amount": employer_amount,
                "deducted_payroll_run_id": PAYROLL_RUN_ID,
                "deducted_payroll_period": PAYROLL_PERIOD,
                "deduction_reason": "正常扣款",
                "status": "deducted",
            })


def _extract_tax_ledgers(wb, records, employees_by_name):
    existing = set()
    for company_id, sheet_name in TAX_SHEETS.items():
        ws = wb[sheet_name]
        company_employees = _employees_for_company(records, company_id)
        for row_offset, row in enumerate(_rows(ws, 1)):
            employee = _mock_aligned_employee(
                records,
                employees_by_name,
                _str(row.get("姓名")),
                company_id=company_id,
                source=sheet_name,
                row_offset=row_offset,
                company_employees=company_employees,
            )
            if not employee:
                continue
            key = employee["employee_id"]
            if key in existing:
                continue
            existing.add(key)
            records["TaxLedger"].append({
                "tax_ledger_id": f"TAX_{TAX_PERIOD.replace('-', '')}_{employee['employee_id']}",
                "employee_id": employee["employee_id"],
                "payroll_run_id": PAYROLL_RUN_ID,
                "tax_period": TAX_PERIOD,
                "tax_entity_company_id": company_id,
                "current_income": _num(row.get("本期收入")),
                "cumulative_income": _num(row.get("累计收入额")),
                "cumulative_deductions": _num(row.get("累计专项扣除")),
                "cumulative_basic_deduction": _num(row.get("累计减除费用")),
                "cumulative_taxable_income": _num(row.get("累计应纳税所得额")),
                "cumulative_tax_payable": _num(row.get("累计应纳税额")),
                "previous_cumulative_tax_withheld": _num(row.get("已缴税额")),
                "current_tax": _num(row.get("应补(退)税额")),
                "source": sheet_name,
                "reconciliation_status": "未核对",
            })


def _extract_payslips(wb, records, employees_by_name):
    ws = wb["202604钉钉工资卡"]
    for row in _rows(ws, 1, max_blank_rows=300):
        employee = employees_by_name.get(_str(row.get("姓名")))
        if not employee:
            continue
        employee_id = employee["employee_id"]
        relationship = _employment_relationship_for_employee(records, employee_id)
        records["Payslip"].append({
            "payslip_id": f"PS_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}",
            "payroll_run_id": PAYROLL_RUN_ID,
            "payroll_line_id": f"PL_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}",
            "employee_id": employee_id,
            "gross_pay_before_deduction": _num(row.get("实际应发工资")),
            "social_security_deduction": _num(row.get("社保合计")),
            "housing_fund_deduction": _num(row.get("住房公积金")),
            "personal_income_tax": _num(row.get("个人所得税")),
            "net_pay": _num(row.get("实发金额")),
            "deduction_periods": CONTRIBUTION_PERIOD,
            "deduction_company_ids": relationship.get("company_id", "") if relationship else "",
            "deduction_note": "由202604钉钉工资卡抽取",
        })


def _extract_cost_entries(wb, records, employees_by_name):
    ws = wb["202604月成本汇总"]
    fields = [
        ("月工资总额（应发）", "扣除前应发工资"),
        ("社保费用（公司）", "公司社保"),
        ("社保费用（个人）", "个人社保"),
        ("公积金费用（公司）", "公司公积金"),
        ("公积金费用（个人）", "个人公积金"),
        ("个人所得税", "个税"),
        ("其他", "其他费用"),
    ]
    seq = 1
    for row in _rows(ws, 1):
        employee = employees_by_name.get(_str(row.get("姓名")))
        if not employee:
            continue
        relationship = _employment_relationship_for_employee(records, employee["employee_id"])
        company_id = relationship.get("company_id", "") if relationship else ""
        affiliation_id = relationship.get("internal_affiliation_id", "") if relationship else ""
        for excel_field, cost_type in fields:
            amount = _num(row.get(excel_field))
            if amount in ("", 0, 0.0):
                continue
            records["CostEntry"].append({
                "cost_entry_id": f"COST_{PAYROLL_PERIOD.replace('-', '')}_{seq:05d}",
                "payroll_run_id": PAYROLL_RUN_ID,
                "employee_id": employee["employee_id"],
                "person_id": employee.get("person_id", ""),
                "consultant_engagement_id": "",
                "employee_snapshot_id": _employee_snapshot_id(employee["employee_id"]),
                "employment_relationship_id": _employment_relationship_id(employee),
                "cost_company_id": company_id,
                "internal_affiliation_id": affiliation_id,
                "cost_period": PAYROLL_PERIOD,
                "payroll_period": PAYROLL_PERIOD,
                "contribution_period": CONTRIBUTION_PERIOD if "社保" in cost_type or "公积金" in cost_type else "",
                "cost_type": cost_type,
                "amount": amount,
                "source_record_type": "202604月成本汇总",
                "source_record_id": f"ROW{row.get('_row_index')}",
            })
            seq += 1


def _extract_consultants(wb, records):
    sheet_name = "马骏老师202604"
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for index, row in enumerate(_rows(ws, 2), start=1):
        name = _str(row.get("姓名"))
        if not name or name in {"总经理：                      财务负责人：                      制表人："}:
            continue
        person_id = f"PER_CONS{index:03d}"
        consultant_engagement_id = f"CENG{index:03d}"
        company_id, company_name = _company(_str(row.get("公司")))
        _ensure_company(records, company_id, company_name)
        records["Person"].append({
            "person_id": person_id,
            "name": name,
            "id_type": "",
            "id_number": "",
            "phone": "",
            "bank_account": "",
            "status": "有效",
        })
        records["ConsultantEngagement"].append({
            "consultant_engagement_id": consultant_engagement_id,
            "person_id": person_id,
            "company_id": company_id,
            "service_start_date": _date(row.get("起薪时间")),
            "service_end_date": "",
            "engagement_type": "顾问",
            "status": "有效",
        })
        settlement_id = f"CFS_{PAYROLL_PERIOD.replace('-', '')}_{consultant_engagement_id}"
        records["ConsultantFeeSettlement"].append({
            "settlement_id": settlement_id,
            "consultant_engagement_id": consultant_engagement_id,
            "person_id": person_id,
            "settlement_period": PAYROLL_PERIOD,
            "company_id": company_id,
            "gross_service_fee": _num(row.get("含税劳务金额")),
            "net_of_tax_fee": _num(row.get("不含税劳务金额")),
            "taxable_income": _num(row.get("应纳税所得额")),
            "withheld_tax": _num(row.get("代扣劳务个税") or row.get("税额")),
            "net_payment": _num(row.get("实发金额")),
            "tax_method": "劳务报酬",
            "status": "calculated",
        })
        records["CostEntry"].append({
            "cost_entry_id": f"COST_{PAYROLL_PERIOD.replace('-', '')}_CONS_{index:03d}",
            "payroll_run_id": PAYROLL_RUN_ID,
            "employee_id": "",
            "person_id": person_id,
            "consultant_engagement_id": consultant_engagement_id,
            "employee_snapshot_id": "",
            "employment_relationship_id": "",
            "cost_company_id": company_id,
            "internal_affiliation_id": "",
            "cost_period": PAYROLL_PERIOD,
            "payroll_period": "",
            "contribution_period": "",
            "cost_type": "顾问劳务费",
            "amount": _num(row.get("含税劳务金额")),
            "source_record_type": "ConsultantFeeSettlement",
            "source_record_id": settlement_id,
        })


def _seed_rules(records):
    split_rules = [
        ("南京大学", 1.0, 0.0),
        ("BOSS", 1.0, 0.0),
        ("实习生", 1.0, 0.0),
        ("人力劳务", 1.0, 0.0),
        ("部门负责人", 0.6, 0.4),
        ("员工", 0.7, 0.3),
    ]
    for idx, (position, basic, perf) in enumerate(split_rules, start=1):
        records["SalarySplitRule"].append({
            "split_rule_id": f"SPLIT_{idx:02d}",
            "effective_period": PAYROLL_PERIOD,
            "position_or_type": position,
            "basic_salary_rate": basic,
            "performance_base_rate": perf,
        })

    grade_rules = [
        ("A", 1.5),
        ("B", 1.2),
        ("C", 1.0),
        ("D", 0.8),
        ("E", 0.6),
        ("无等级", 0.0),
    ]
    for idx, (grade, coefficient) in enumerate(grade_rules, start=1):
        records["PerformanceGradeRule"].append({
            "grade_rule_id": f"PERF_RULE_{idx:02d}",
            "effective_period": PAYROLL_PERIOD,
            "performance_grade": grade,
            "coefficient": coefficient,
        })

    social_rules = [
        ("COMP_SCNY", 0.0032),
        ("COMP_NDSC", 0.0055),
        ("COMP_JS", 0.002),
    ]
    for company_id, injury_rate in social_rules:
        records["SocialInsuranceRule"].append({
            "social_rule_id": f"SOC_RULE_{company_id}",
            "company_id": company_id,
            "effective_period": CONTRIBUTION_PERIOD,
            "lower_bound": 4952,
            "upper_bound": 24762,
            "employer_pension_rate": 0.16,
            "employer_injury_rate": injury_rate,
            "employer_unemployment_rate": 0.005,
            "employer_medical_rate": 0.07,
            "employer_maternity_rate": 0.008,
            "personal_pension_rate": 0.08,
            "personal_unemployment_rate": 0.005,
            "personal_medical_rate": 0.02,
            "personal_serious_illness_amount": 10,
        })

    for company_id, rate in [("COMP_SCNY", 0.10), ("COMP_NDSC", 0.10), ("COMP_JS", 0.12)]:
        records["HousingFundRule"].append({
            "housing_rule_id": f"HF_RULE_{company_id}",
            "company_id": company_id,
            "effective_period": CONTRIBUTION_PERIOD,
            "employer_rate": rate,
            "personal_rate": rate,
            "rounding_rule": "ROUND0",
        })

    tax_rules = [
        (0, 36000, 0.03, 0),
        (36000, 144000, 0.10, 2520),
        (144000, 300000, 0.20, 16920),
        (300000, 420000, 0.25, 31920),
        (420000, 660000, 0.30, 52920),
        (660000, 960000, 0.35, 85920),
        (960000, "", 0.45, 181920),
    ]
    for idx, (lower, upper, rate, quick) in enumerate(tax_rules, start=1):
        records["TaxRateRule"].append({
            "tax_rule_id": f"TAX_RATE_{idx:02d}",
            "lower_bound": lower,
            "upper_bound": upper,
            "tax_rate": rate,
            "quick_deduction": quick,
        })


def _write_csvs(data_dir: Path, objects: dict[str, list[str]], records: dict[str, list[dict]]):
    data_dir.mkdir(parents=True, exist_ok=True)
    for object_type, fields in objects.items():
        path = data_dir / f"{_snake(object_type)}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for row in records.get(object_type, []):
                writer.writerow({field: _csv_value(row.get(field, "")) for field in fields})


def _write_benchmark_csvs(
    benchmark_dir: Path,
    objects: dict[str, list[str]],
    records: dict[str, list[dict]],
):
    benchmark_dir.mkdir(parents=True, exist_ok=True)
    for object_type in sorted(BENCHMARK_OBJECTS):
        fields = objects.get(object_type)
        if not fields:
            continue
        path = benchmark_dir / f"{_snake(object_type)}_expected.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for row in records.get(object_type, []):
                writer.writerow({field: _csv_value(row.get(field, "")) for field in fields})


def _clear_generated_outputs(records: dict[str, list[dict]]):
    for object_type in GENERATED_OUTPUT_OBJECTS:
        records[object_type] = []


def _convert_tax_ledgers_to_historical_inputs(records: dict[str, list[dict]]):
    deductions_by_employee: dict[str, dict[str, float]] = {}
    for deduction in records.get("ContributionDeduction", []):
        employee_id = deduction.get("employee_id", "")
        deduction_type = deduction.get("deduction_type", "")
        deductions_by_employee.setdefault(employee_id, {"社保": 0.0, "公积金": 0.0})
        if deduction_type in deductions_by_employee[employee_id]:
            deductions_by_employee[employee_id][deduction_type] += _as_float(
                deduction.get("personal_amount")
            )

    for ledger in records.get("TaxLedger", []):
        employee_id = ledger.get("employee_id", "")
        current_income = _as_float(ledger.get("current_income"))
        current_social = deductions_by_employee.get(employee_id, {}).get("社保", 0.0)
        current_housing = deductions_by_employee.get(employee_id, {}).get("公积金", 0.0)
        ledger["current_income"] = ""
        ledger["cumulative_income"] = max(
            round(_as_float(ledger.get("cumulative_income")) - current_income, 2),
            0.0,
        )
        ledger["cumulative_deductions"] = max(
            round(
                _as_float(ledger.get("cumulative_deductions"))
                - current_social
                - current_housing,
                2,
            ),
            0.0,
        )
        ledger["cumulative_taxable_income"] = ""
        ledger["cumulative_tax_payable"] = ""
        ledger["current_tax"] = ""
        ledger["source"] = "historical_input_from_excel"
        ledger["reconciliation_status"] = "待计算"


def _rows(ws, header_row: int, max_blank_rows: int = 50):
    iterator = ws.iter_rows(min_row=header_row, values_only=True)
    try:
        header_values = next(iterator)
    except StopIteration:
        return
    headers = [_str(value) for value in header_values]
    blank_rows = 0
    for row_index, values in enumerate(iterator, start=header_row + 1):
        row = {"_row_index": row_index}
        any_value = False
        for value, header in zip(values, headers):
            if not header:
                continue
            if value not in (None, ""):
                any_value = True
            row[header] = value
        if any_value:
            blank_rows = 0
            yield row
        else:
            blank_rows += 1
            if blank_rows >= max_blank_rows:
                break


def _ensure_company(records, company_id: str, company_name: str):
    if any(row.get("company_id") == company_id for row in records["Company"]):
        return
    records["Company"].append({
        "company_id": company_id,
        "company_name": company_name,
        "payroll_enabled": 1,
        "tax_entity_name": company_name,
        "social_security_account": "",
        "housing_fund_account": "",
    })


def _employees_for_company(records, company_id: str) -> list[dict]:
    employee_ids = {
        relationship.get("employee_id")
        for relationship in records["EmploymentRelationship"]
        if relationship.get("company_id") == company_id
    }
    return [
        employee for employee in records["Employee"]
        if employee.get("employee_id") in employee_ids
        and employee.get("status") != "来源表补充"
    ]


def _mock_aligned_employee(
    records,
    employees_by_name: dict[str, dict],
    source_name: str,
    company_id: str,
    source: str,
    row_offset: int,
    company_employees: list[dict],
) -> dict | None:
    source_name = _str(source_name)
    if not source_name:
        return None

    direct = employees_by_name.get(source_name)
    direct_relationship = (
        _employment_relationship_for_employee(records, direct.get("employee_id", ""))
        if direct else None
    )
    if direct_relationship and direct_relationship.get("company_id") == company_id:
        return direct

    if row_offset < len(company_employees):
        employee = company_employees[row_offset]
        person = _person_for_employee(records, employee)
        employee_name = person.get("name", "") if person else ""
        _append_mock_warning(
            records,
            target_type="Employee",
            target_id=employee["employee_id"],
            rule_code="mock_source_name_aligned_by_company_order",
            message=(
                f"{source} 第 {row_offset + 1} 条来源姓名“{source_name}”"
                f"已按公司内行序对齐到员工 {employee['employee_id']}（{employee_name}）"
            ),
        )
        return employee

    _append_mock_warning(
        records,
        target_type="Employee",
        target_id=f"{company_id}_ROW_{row_offset + 1}",
        rule_code="mock_source_row_overflow",
        message=(
            f"{source} 第 {row_offset + 1} 条来源姓名“{source_name}”"
            f"超出公司 {company_id} 的 canonical 员工列表，已跳过"
        ),
    )
    return None


def _append_mock_warning(
    records,
    target_type: str,
    target_id: str,
    rule_code: str,
    message: str,
):
    validation_id = f"VAL_{len(records['PayrollValidationResult']) + 1:04d}"
    records["PayrollValidationResult"].append({
        "validation_id": validation_id,
        "payroll_run_id": PAYROLL_RUN_ID,
        "severity": "warning",
        "rule_code": rule_code,
        "target_type": target_type,
        "target_id": target_id,
        "message": message,
        "status": "open",
    })


def _get_or_create_employee(
    records,
    employees_by_name: dict[str, dict],
    name: str,
    company_id: str,
    source: str,
    salary_start_date: str = "",
    id_type: str = "",
    id_number: str = "",
) -> dict | None:
    name = _str(name)
    if not name:
        return None

    employee = employees_by_name.get(name)
    if employee:
        person = _person_for_employee(records, employee)
        if person:
            if id_type and not person.get("id_type"):
                person["id_type"] = id_type
            if id_number and not person.get("id_number"):
                person["id_number"] = id_number
        return employee

    employee_id = f"EMP{len(records['Employee']) + 1:03d}"
    person_id = f"PER{len(records['Person']) + 1:03d}"
    employee = {
        "employee_id": employee_id,
        "person_id": person_id,
        "employee_number": "",
        "department": "",
        "position": "",
        "employment_type": "",
        "hire_date": "",
        "salary_start_date": salary_start_date,
        "termination_date": "",
        "status": "来源表补充",
    }
    records["Person"].append({
        "person_id": person_id,
        "name": name,
        "id_type": id_type,
        "id_number": id_number,
        "phone": "",
        "bank_account": "",
        "status": "有效",
    })
    records["Employee"].append(employee)
    employees_by_name[name] = employee
    records["EmploymentRelationship"].append({
        "employment_relationship_id": _employment_relationship_id(employee),
        "employee_id": employee_id,
        "person_id": person_id,
        "effective_date": salary_start_date or f"{PAYROLL_PERIOD}-01",
        "expiry_date": "",
        "company_id": company_id,
        "internal_affiliation_id": "",
        "relationship_type": "来源表补充",
        "relationship_reason": "来源表补充",
        "source": source,
        "notes": "由非总工资表来源补充生成的任职关系记录",
    })
    records["PayrollValidationResult"].append({
        "validation_id": f"VAL_EMP_SOURCE_{employee_id}",
        "payroll_run_id": PAYROLL_RUN_ID,
        "severity": "warning",
        "rule_code": "employee_from_secondary_source",
        "target_type": "Employee",
        "target_id": employee_id,
        "message": f"员工由 {source} 补充生成，未在总工资表中按姓名匹配到",
        "status": "open",
    })
    return employee


def _company(raw: str) -> tuple[str, str]:
    text = _str(raw)
    if text in COMPANY_ALIASES:
        return COMPANY_ALIASES[text]
    normalized = re.sub(r"[（）()\\s]", "", text)
    for key, value in COMPANY_ALIASES.items():
        if normalized and re.sub(r"[（）()\\s]", "", key) == normalized:
            return value
    slug = re.sub(r"\\W+", "_", text).strip("_") or "UNKNOWN"
    return f"COMP_{slug}", text


def _company_name(company_id: str) -> str:
    for cid, name in COMPANY_ALIASES.values():
        if cid == company_id:
            return name
    return company_id


def _affiliation_id(raw_company: str, company_id: str) -> str:
    text = _str(raw_company)
    if "（" in text or "(" in text:
        slug = re.sub(r"\\W+", "_", text).strip("_")
        return f"AFF_{company_id}_{slug}"
    return ""


def _num_by_prefix(row: dict, prefix: str):
    for key, value in row.items():
        if key.startswith(prefix):
            return _num(value)
    return ""


def _employment_relationship_id(employee: dict) -> str:
    employee_id = employee.get("employee_id", "")
    if not employee_id:
        return ""
    return f"EREL_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}"


def _employment_relationship_for_employee(records, employee_id: str) -> dict | None:
    for relationship in records["EmploymentRelationship"]:
        if relationship.get("employee_id") == employee_id:
            return relationship
    return None


def _person_for_employee(records, employee: dict) -> dict | None:
    person_id = employee.get("person_id", "")
    for person in records["Person"]:
        if person.get("person_id") == person_id:
            return person
    return None


def _employee_snapshot_id(employee_id: str) -> str:
    if not employee_id:
        return ""
    return f"ESNAP_{PAYROLL_PERIOD.replace('-', '')}_{employee_id}"


def _num(value: Any):
    if value in (None, ""):
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return ""


def _as_float(value: Any) -> float:
    if value in (None, ""):
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def _date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _str(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _snake(name: str) -> str:
    result = []
    for i, ch in enumerate(name):
        if ch.isupper() and i > 0:
            result.append("_")
        result.append(ch.lower())
    return "".join(result)


if __name__ == "__main__":
    raise SystemExit(main())
